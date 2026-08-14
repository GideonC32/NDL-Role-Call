import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import sqlite3
import datetime
import os

DB_PATH = '/workspace/ndl_monitoring.db'

def generate_excel_report(output_path):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 1. Fetch learners and their latest update
    cursor.execute('''
    SELECT 
        l.name, 
        l.assigned_teacher, 
        l.room,
        u.status,
        u.event,
        u.recorded_by,
        u.date,
        u.time
    FROM learners l
    LEFT JOIN (
        SELECT u1.* FROM updates u1
        JOIN (
            SELECT learner_id, MAX(timestamp) as max_ts 
            FROM updates 
            GROUP BY learner_id
        ) u2 ON u1.learner_id = u2.learner_id AND u1.timestamp = u2.max_ts
    ) u ON l.id = u.learner_id
    ORDER BY l.name ASC
    ''')
    learners_status = cursor.fetchall()
    
    # 2. Fetch full historical updates
    cursor.execute('''
    SELECT timestamp, learner_name, assigned_teacher, recorded_by, event, status, note, date, time
    FROM updates
    ORDER BY timestamp DESC
    ''')
    history_logs = cursor.fetchall()
    
    conn.close()
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Setup styles
    font_title = Font(name='Calibri', size=16, bold=True, color='1F4E78')
    font_section = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_bold = Font(name='Calibri', size=11, bold=True)
    font_normal = Font(name='Calibri', size=11)
    font_italic = Font(name='Calibri', size=9, italic=True, color='595959')
    
    fill_header = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    fill_section = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    fill_zebra = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_green_status = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Present
    fill_red_status = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # Absent
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    border_medium = Side(border_style="medium", color="2F5496")
    border_double = Side(border_style="double", color="2F5496")
    
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_header_bottom = Border(bottom=border_medium)
    
    # -------------------------------------------------------------
    # SHEET 1: Current Status
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Current Status"
    ws1.sheet_view.showGridLines = True
    
    # Title Block
    ws1['A1'] = "NDL Learner Monitoring & Roll-Call App — Current Status"
    ws1['A1'].font = font_title
    
    today_str = datetime.date.today().strftime('%Y-%m-%d')
    ws1['A2'] = f"Generated: {today_str} | v1.0 (Live Shared Database Report)"
    ws1['A2'].font = font_italic
    
    # Summary Dashboard Block
    ws1['A4'] = "Metric Summary"
    ws1['B4'] = "Value"
    ws1['A4'].font = font_section
    ws1['B4'].font = font_section
    ws1['A4'].fill = fill_section
    ws1['B4'].fill = fill_section
    ws1['A4'].border = border_cell
    ws1['B4'].border = border_cell
    
    # Our data starts at row 11 and ends at row 26 (16 learners)
    summary_metrics = [
        ("Total Learners", "=COUNTA(A11:A26)"),
        ("Present", '=COUNTIF(D11:D26, "*Present")'),
        ("Absent", '=COUNTIF(D11:D26, "*Absent")'),
        ("No Update Yet", '=COUNTIF(D11:D26, "*No update yet")')
    ]
    
    for idx, (metric, formula) in enumerate(summary_metrics):
        r = 5 + idx # rows 5, 6, 7, 8
        ws1.cell(row=r, column=1, value=metric).font = font_bold
        ws1.cell(row=r, column=1).border = border_cell
        ws1.cell(row=r, column=2, value=formula).font = font_normal
        ws1.cell(row=r, column=2).alignment = Alignment(horizontal='right')
        ws1.cell(row=r, column=2).border = border_cell
        if metric == "Present":
            ws1.cell(row=r, column=2).fill = fill_green_status
        elif metric == "Absent":
            ws1.cell(row=r, column=2).fill = fill_red_status
        else:
            ws1.cell(row=r, column=2).fill = fill_white
            
    # Header Row for Table (Row 10)
    headers = ["Learner Name", "Assigned Teacher", "Room", "Latest Status", "Latest Event", "Updated By", "Date", "Time"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=10, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center' if col_idx > 3 else 'left')
        cell.border = border_header_bottom
        
    # Write Data (Rows 11 to 26)
    for row_idx, data in enumerate(learners_status, 11):
        name, teacher, room, status, event, recorded_by, d_val, t_val = data
        
        status_display = "⚪ No update yet" if not status else (f"🟢 {status}")
        event_display = event if event else "N/A"
        recorded_display = recorded_by if recorded_by else "N/A"
        date_display = d_val if d_val else "N/A"
        time_display = t_val if t_val else "N/A"
        
        row_values = [name, teacher, room, status_display, event_display, recorded_display, date_display, time_display]
        row_fill = fill_zebra if row_idx % 2 == 0 else fill_white
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_normal
            cell.border = border_cell
            cell.fill = row_fill
            
            if col_idx in [3, 4, 7, 8]: # Room, Status, Date, Time
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
                
            if col_idx == 4: # Status Column
                if "Present" in val:
                    cell.fill = fill_green_status
                    cell.font = font_bold
                elif "Absent" in val:
                    cell.fill = fill_red_status
                    cell.font = font_bold
                    
    # Auto-fit column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
    ws1.column_dimensions['A'].width = 28 # Learner Name wider
    
    # Freeze pane on row 11 (so header row 10 is frozen)
    ws1.freeze_panes = 'A11'
    
    # -------------------------------------------------------------
    # SHEET 2: History Log
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Historical Log")
    
    # Title Block
    ws2['A1'] = "NDL Learner Monitoring & Roll-Call App — Full History Log"
    ws2['A1'].font = font_title
    ws2['A2'] = f"Generated: {today_str} | Contains all chronological safety updates"
    ws2['A2'].font = font_italic
    
    # Header Row
    headers_h = ["System Timestamp", "Learner Name", "Assigned Teacher", "Recorded By", "Event", "Attendance Status", "Optional Note", "Date", "Time"]
    for col_idx, h in enumerate(headers_h, 1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center' if col_idx in [1, 6, 8, 9] else 'left')
        cell.border = border_header_bottom
        
    # Write History
    if not history_logs:
        ws2.cell(row=5, column=1, value="No updates have been recorded yet in this session.").font = font_italic
        ws2.merge_cells('A5:I5')
    else:
        for row_idx, h_data in enumerate(history_logs, 5):
            ts, name, teacher, rec_by, event, status, note, d_val, t_val = h_data
            
            row_values = [ts, name, teacher, rec_by, event, status, note if note else "", d_val, t_val]
            row_fill = fill_zebra if row_idx % 2 == 0 else fill_white
            
            for col_idx, val in enumerate(row_values, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_normal
                cell.border = border_cell
                cell.fill = row_fill
                
                if col_idx in [1, 6, 8, 9]:
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')
                    
                if col_idx == 6:
                    if val == "Present":
                        cell.fill = fill_green_status
                        cell.font = font_bold
                    elif val == "Absent":
                        cell.fill = fill_red_status
                        cell.font = font_bold

    # Auto-fit columns
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
    ws2.column_dimensions['B'].width = 28 # Learner Name
    ws2.column_dimensions['A'].width = 22 # Timestamp
    ws2.column_dimensions['G'].width = 25 # Note
    
    # Freeze pane
    ws2.freeze_panes = 'A5'
    
    # Save the file
    wb.save(output_path)
